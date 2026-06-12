#!/usr/bin/env python3
"""
format_xlsx.py — rend le CSV en classeur Excel mis en forme (skill format-xlsx).

Le CSV reste la BASE DE DONNEES canonique (deterministe, dedup, diff entre scans).
Ce script en derive un .xlsx LISIBLE :
  - en-tete fige + filtres automatiques + volets geles,
  - couleur de ligne par bucket (finance quant en avant),
  - badge couleur sur in_europe (yes vert / no rouge / ? orange),
  - colonne url cliquable (hyperlien),
  - onglet "Toutes" + onglets "Europe" et "Finance quant" + onglet "Resume".

Aucun reseau. Lecture CSV -> ecriture XLSX via openpyxl.
Usage: python format_xlsx.py <stages.csv> <sortie.xlsx>
"""
import csv, argparse, os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["company", "title", "location", "in_europe", "bucket",
           "source", "url", "first_seen", "last_seen"]
NICE = {"company": "Entreprise", "title": "Poste", "location": "Localisation",
        "in_europe": "Europe", "bucket": "Categorie", "source": "Source",
        "url": "Lien", "first_seen": "Vue le", "last_seen": "Revue le"}

BUCKET_FILL = {"bank_quant": "D9E1F2", "hedge_fund_quant": "DDEBF7",
               "data_scientist": "E2EFDA", "data_science_ai": "FCE4D6",
               "data_analyst": "FFF2CC", "consulting_data": "EDEDED"}
BUCKET_LABEL = {"bank_quant": "Quant - Banque", "hedge_fund_quant": "Quant - Hedge fund",
                "data_scientist": "Data Scientist", "data_science_ai": "Data Science / IA",
                "data_analyst": "Data Analyst", "consulting_data": "Consulting Data", "?": "A classer"}
EUROPE_FILL = {"yes": "C6EFCE", "no": "FFC7CE", "?": "FFEB9C"}
EUROPE_FONT = {"yes": "006100", "no": "9C0006", "?": "9C6500"}

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LINK_FONT = Font(color="0563C1", underline="single")


def read_rows(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def style_sheet(ws, rows, heading):
    ws.append([f"{heading}  -  {len(rows)} offres"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.cell(1, 1).font = Font(bold=True, size=13, color="1F3864")
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.append([NICE[h] for h in HEADERS])
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(2, c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for r in rows:
        bucket = r.get("bucket", "?")
        eu = r.get("in_europe", "?")
        vals = [r.get("company", ""), r.get("title", ""), r.get("location", ""),
                eu, BUCKET_LABEL.get(bucket, bucket), r.get("source", ""),
                r.get("url", ""), r.get("first_seen", ""), r.get("last_seen", "")]
        ws.append(vals)
        row_i = ws.max_row
        fill = BUCKET_FILL.get(bucket)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row_i, c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (2, 3)))
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
        ecell = ws.cell(row_i, 4)
        ecell.fill = PatternFill("solid", fgColor=EUROPE_FILL.get(eu, "FFFFFF"))
        ecell.font = Font(bold=True, color=EUROPE_FONT.get(eu, "000000"))
        ecell.alignment = Alignment(horizontal="center", vertical="center")
        url = r.get("url", "")
        if url.startswith("http"):
            lcell = ws.cell(row_i, 7)
            lcell.hyperlink = url
            lcell.font = LINK_FONT
            lcell.value = "ouvrir l'offre"
    widths = [22, 42, 24, 9, 18, 18, 16, 12, 12]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{ws.max_row}"


def _head(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c); cell.fill = HEAD_FILL; cell.font = HEAD_FONT


def summary_sheet(ws, rows):
    ws.append(["Resume du scan"]); ws.cell(1, 1).font = Font(bold=True, size=14, color="1F3864")
    ws.append([])
    bc = Counter(r.get("bucket", "?") for r in rows)
    ws.append(["Par categorie", "Nb"]); _head(ws, ws.max_row, 2)
    order = ["bank_quant", "hedge_fund_quant", "data_scientist", "data_science_ai",
             "data_analyst", "consulting_data", "?"]
    for b in order:
        if bc.get(b):
            ws.append([BUCKET_LABEL.get(b, b), bc[b]])
            fill = BUCKET_FILL.get(b)
            if fill:
                ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=fill)
    ws.append([])
    ec = Counter(r.get("in_europe", "?") for r in rows)
    ws.append(["Par localisation", "Nb"]); _head(ws, ws.max_row, 2)
    for k, lab in [("yes", "Europe (oui)"), ("?", "A confirmer"), ("no", "Hors Europe")]:
        ws.append([lab, ec.get(k, 0)])
        ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=EUROPE_FILL.get(k))
    ws.append([]); ws.append(["Total offres", len(rows)])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 10


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("csv"); ap.add_argument("xlsx")
    a = ap.parse_args()
    if not os.path.exists(a.csv):
        print(f"[format] ERREUR: CSV introuvable: {a.csv}\n"
              "  -> lance d'abord un scan ; travaille dans un dossier connecte a Cowork "
              "ou le CSV existe.", file=__import__("sys").stderr)
        raise SystemExit(2)
    rows = read_rows(a.csv)
    wb = Workbook()
    summary_sheet(wb.active, rows)
    wb.active.title = "Resume"
    style_sheet(wb.create_sheet("Toutes"), rows, "Toutes les offres")
    eu = [r for r in rows if r.get("in_europe") in ("yes", "?")]
    style_sheet(wb.create_sheet("Europe"), eu, "Europe (priorite)")
    fin = [r for r in rows if r.get("bucket") in ("bank_quant", "hedge_fund_quant")]
    style_sheet(wb.create_sheet("Finance quant"), fin, "Finance quant")
    os.makedirs(os.path.dirname(os.path.abspath(a.xlsx)) or ".", exist_ok=True)
    wb.save(a.xlsx)
    print(f"[format] {len(rows)} offres -> {a.xlsx} (Toutes / Europe {len(eu)} / Finance quant {len(fin)} / Resume)")


if __name__ == "__main__":
    main()
